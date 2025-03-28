from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Puesto, Area, Worker, Marca, Category, Processor, Ram, Disco, Warranty, Mouse, Keyboard, Monitor, Camera, Audifono, Stabilizer, Supresor, Cooler, Device

# Create your views here.
def usuario(request):
  search_term = request.GET.get('search', '')
  workers = Worker.objects.filter(name__contains=search_term)
  return render(request, 'devices/inventario.html', {'workers': workers, 'search_term': search_term})

def export_excel_workers(request):
  workers = Worker.objects.filter(name__contains=request.GET.get('search', ''))
  wb = Workbook()
  ws = wb.active
  ws.title = 'Usuarios'

  ws.merge_cells('A1:F1')
  ws['A1'] = 'USUARIOS'
  ws['A1'].font = Font(color=Color('FFFFFF'), bold=True, size=14)
  ws['A1'].fill = PatternFill(fgColor='2DADE6', fill_type='solid')
  ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

  ws['A2'] = 'Usuario'
  ws['B2'] = 'Area'
  ws['C2'] = 'Puesto'
  ws['D2'] = 'Activo'
  ws['E2'] = 'Fecha-Creación'
  ws['F2'] = 'Fecha-Actualización'

  font_style_name = Font(color=Color('FFFFFF'), bold=True)
  fill_style = PatternFill(fgColor='2BB2C7', fill_type='solid')
  alignment_style = Alignment(horizontal='center', vertical='center')
  border_style = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

  ws['A2'].font = font_style_name
  ws['A2'].fill = fill_style
  ws['A2'].alignment = alignment_style
  ws['A2'].border = border_style
  ws['B2'].font = font_style_name
  ws['B2'].fill = fill_style
  ws['B2'].alignment = alignment_style
  ws['B2'].border = border_style
  ws['C2'].font = font_style_name
  ws['C2'].fill = fill_style
  ws['C2'].alignment = alignment_style
  ws['C2'].border = border_style
  ws['D2'].font = font_style_name
  ws['D2'].fill = fill_style
  ws['D2'].alignment = alignment_style
  ws['D2'].border = border_style
  ws['E2'].font = font_style_name
  ws['E2'].fill = fill_style
  ws['E2'].alignment = alignment_style
  ws['E2'].border = border_style
  ws['F2'].font = font_style_name
  ws['F2'].fill = fill_style
  ws['F2'].alignment = alignment_style
  ws['F2'].border = border_style

  ws.column_dimensions[get_column_letter(1)].width = 30
  ws.column_dimensions[get_column_letter(2)].width = 37
  ws.column_dimensions[get_column_letter(3)].width = 30
  ws.column_dimensions[get_column_letter(4)].width = 7
  ws.column_dimensions[get_column_letter(5)].width = 20
  ws.column_dimensions[get_column_letter(6)].width = 20

  row=3

  for user in workers:
    ws.cell(row=row, column=1, value=user.name)
    ws.cell(row=row, column=2, value=', '.join([area.name for area in user.area.all()]))
    ws.cell(row=row, column=3, value=', '.join([puesto.name for puesto in user.puesto.all()]))
    ws.cell(row=row, column=4, value=user.active)
    ws.cell(row=row, column=5, value=user.created.strftime('%d/%m/%Y'))
    ws.cell(row=row, column=6, value=user.updated.strftime('%d/%m/%Y'))

    ws.cell(row=row, column=4).alignment = alignment_style
    ws.cell(row=row, column=5).alignment = alignment_style
    ws.cell(row=row, column=6).alignment = alignment_style

    ws.cell(row=row, column=1).border = border_style
    ws.cell(row=row, column=2).border = border_style
    ws.cell(row=row, column=3).border = border_style
    ws.cell(row=row, column=4).border = border_style
    ws.cell(row=row, column=5).border = border_style
    ws.cell(row=row, column=6).border = border_style

    row += 1

  response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheeetml.sheet')
  response['Content-Disposition'] = 'attachment; filename=Usuarios de INEL.xlsx'
  wb.save(response)
  
  return response

def computadora(request):
  computadoras = Device.objects.all()
  return render(request, 'devices/computadora.html', {'computadoras': computadoras})

def mouse(request):
  mouses = Mouse.objects.all()
  return render(request, 'devices/mouse.html', {'mouses': mouses})

def teclado(request):
  teclados = Keyboard.objects.all()
  return render(request, 'devices/teclado.html', {'teclados': teclados})

def monitor(request):
  monitores = Monitor.objects.all()
  return render(request, 'devices/monitor.html', {'monitores': monitores})

def camara(request):
  camaras = Camera.objects.all()
  return render(request, 'devices/camara.html', {'camaras': camaras})

def audifono(request):
  audifonos = Audifono.objects.all()
  return render(request, 'devices/audifono.html', {'audifonos': audifonos})

def estabilizador(request):
  estabilizadores = Stabilizer.objects.all()
  return render(request, 'devices/estabilizador.html', {'estabilizadores': estabilizadores})

def cooler(request):
  coolers = Cooler.objects.all()
  return render(request, 'devices/cooler.html', {'coolers': coolers})

def supresor(request):
  supresores = Supresor.objects.all()
  return render(request, 'devices/supresor.html', {'supresores': supresores})
